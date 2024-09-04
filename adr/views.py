from django.shortcuts import get_object_or_404, render # type: ignore
import openpyxl # type: ignore # Importamos openpyxl para trabajar con archivos de Excel
from openpyxl import Workbook # type: ignore # Importamos Workbook para trabajar con archivos de Excel
from openpyxl.styles import Font, PatternFill, Border, Side # type: ignore
from openpyxl.utils import get_column_letter # type: ignore # Importa get_column_letter para obtener la letra de la columna
from openpyxl.styles import Font # type: ignore # Importa Font para trabajar con la fuente
from datetime import datetime  # Importa datetime correctamente
from django.utils.timezone import now # type: ignore # Importa now para el tiempo actual
from django.http import HttpResponse # type: ignore # Importa HttpResponse para devolver una respuesta HTTP
from django.views import View # type: ignore # Importamos View para trabajar con las vistas URLs
from django.views.generic import ListView, TemplateView, CreateView, UpdateView, DeleteView, DetailView # type: ignore # Importamos ListView, TemplateView, CreateView, UpdateView, DeleteView, DetailView
from accounts.models import Profile
from django.shortcuts import redirect, render # type: ignore # Importamos render y redirect
from django.urls import reverse_lazy, reverse # type: ignore # Importamos reverse_lazy
from django.contrib.auth.models import Group # type: ignore # Importamos la estructura de los grupos de Django
from django.contrib.auth.models import User # type: ignore # Importamos el modelo de usuario
from django.contrib.auth import authenticate, login, logout # type: ignore # Importamos la autenticación
from .funciones import plural_singular, filtrar_y_paginar # Importa la función plural_singular
from .decorators import add_group_name_to_context, get_group_and_color
from django.contrib.auth.mixins import UserPassesTestMixin, LoginRequiredMixin # type: ignore # Importamos la clase UserPassesTestMixin para proteger las vistas URLs
import os # Importamos la librería os
from django.contrib import messages  # type: ignore # Importamos mensajes
from django.conf import settings    # type: ignore # Importamos la configuración de Django
from django.contrib.auth.views import PasswordChangeView  # type: ignore # Importamos la clase PasswordChangeView
from django.contrib.auth import update_session_auth_hash # type: ignore # Importamos la clase update_session_auth_hash
from django.contrib.auth.views import LoginView # type: ignore # Importamos la clase LoginView
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger # type: ignore # Importamos la clase Paginator para trabajar con la paginación
from .models import AllInOne, AllInOneAdmins, Notebook, MiniPC, Proyectores, BodegaADR, Azotea, Reporte
from .forms import LoginForm, UserCreationForm, ProfileForm, UserForm, AllInOneForm, AllInOneAdminsForm, NotebooksForm, MiniPCForm, ProyectoresForm, BodegaADRForm, AzoteaForm, \
      ReporteForm# Importamos los form para crear los formularios en las vistas

# ____________________________________________________________________________________________________________________________
#TEMPLATES BASADAS EN CLASES PARA EXPORTAR UN ARCHIVO EXCEL

@add_group_name_to_context
class DescargarExcelView(LoginRequiredMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        model_name = kwargs.get('model_name')
        
        # Obtener la fecha actual en formato día-mes-año
        fecha_actual = now().strftime('%d-%m-%Y')
        
        # Verificar qué modelo se está solicitando
        if model_name == 'allinone':
            queryset = AllInOne.objects.all()
            filename = f'AllInOne_{fecha_actual}.xlsx'
        elif model_name == 'allinoneadmin':
            queryset = AllInOneAdmins.objects.all()
            filename = f'AllInOneAdmins_{fecha_actual}.xlsx'
        elif model_name == 'notebook':
            queryset = Notebook.objects.all()
            filename = f'Notebooks_{fecha_actual}.xlsx'
        elif model_name == 'minipc':
            queryset = MiniPC.objects.all()
            filename = f'MiniPCs_{fecha_actual}.xlsx'
        elif model_name == 'proyector':
            queryset = Proyectores.objects.all()
            filename = f'Proyectores_{fecha_actual}.xlsx'
        elif model_name == 'bodegaadr':
            queryset = BodegaADR.objects.all()
            filename = f'BodegaADR_{fecha_actual}.xlsx'
        elif model_name == 'azotea_adr':
            queryset = Azotea.objects.all()
            filename = f'Azotea_{fecha_actual}.xlsx'
        elif model_name == 'reporte':
            queryset = Reporte.objects.all()
            filename = f'Reporte_{fecha_actual}.xlsx'
        else:
            return HttpResponse(status=404)

        # Crear un libro de trabajo y una hoja de trabajo
        wb = Workbook()
        ws = wb.active
        ws.title = model_name.capitalize()

        # Mapear los nombres de las columnas originales a los nombres deseados
        column_names = {
            'id': 'N°',
            'entregado_por': 'ENTREGADO POR',
            'activo': 'ACTIVO',
            'marca': 'MARCA',
            'modelo': 'MODELO',
            'n_serie': 'N° SERIE',
            'bolso': 'BOLSO',
            'equipo': 'EQUIPO',
            'cargador': 'CARGADOR',
            'unive': 'UNIVE',
            'bdo': 'BDO',
            'netbios': 'NETBIOS',
            'ubicacion_all_in_one': 'UBICACIÓN',
            'ubicacion_all_in_one_admin': 'UBICACIÓN',
            'ubicacion_notebook': 'UBICACIÓN',
            'ubicacion_mini_pc': 'UBICACIÓN',
            'ubicacion_proyector': 'UBICACIÓN',
            'ubicacion': 'UBICACIÓN',
            'sala': 'SALA',
            'estado_activo': 'ESTADO',
            'all_one_creado': 'REGISTRADO POR',
            'all_one_adm_creado': 'REGISTRADO POR',
            'notebook_creado': 'REGISTRADO POR',
            'mini_pc_creado': 'REGISTRADO POR',
            'proyector_creado': 'REGISTRADO POR',
            'bodega_adr_creado': 'REGISTRADO POR',
            'azotea_creado': 'REGISTRADO POR',
            'registrador_por': 'REGISTRADOR POR',
            'reporte_creado': 'RECEPCIONADO POR',
            'en_reporte': 'EN REPORTE',
            'operador_entrega': 'OPERADOR ENTREGA',
            'recibido_por': 'RECIBIDO POR',
            'entregado': 'ESTADO REPORTE',
            'fecha': 'FECHA',
            'fecha_recepcion': 'FECHA',
            'fecha_entrega': 'FECHA ENTREGA',
        }

        # Añadir los nombres de las columnas al archivo Excel
        columns = [field.name for field in queryset.model._meta.fields]
        column_headers = [column_names.get(col, col) for col in columns]
        ws.append(column_headers)

        # Estilo para las cabeceras
        bold_font = Font(bold=True, color='FFFFFF')  # Letra blanca
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Fondo rojo
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))  # Bordes delgados

        # Aplicar formato en negrita, color a las cabeceras y bordes
        for cell in ws[1]:
            cell.font = bold_font
            cell.fill = red_fill
            cell.border = border

        # Añadir los datos del queryset al archivo Excel
        for obj in queryset:
            row = []
            for field in columns:
                value = getattr(obj, field)
                if isinstance(value, datetime):
                    # Formatear las fechas
                    value = value.strftime('%d/%m/%Y')
                elif hasattr(value, 'first_name') and hasattr(value, 'last_name'):
                    # Convertir instancias de User a nombres completos
                    value = f"{value.first_name} {value.last_name}"
                elif isinstance(value, bool):
                    # Convertir valores booleanos a 'Sí' o 'No'
                    value = 'Sí' if value else 'No'
                row.append(value)
            ws.append(row)

        # Aplicar bordes a todas las celdas
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border

        #ajustar automáticamente el ancho de las columnas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)  # Obtén la letra de la columna
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Preparar la respuesta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)

        return response

#__________________________________________________________________________________________________________________________

# VISTA BASADA EN CLASES DEL LOGIN
@add_group_name_to_context
class Login(LoginView):
    template_name = 'index.html'
    form_class = LoginForm
    success_url = reverse_lazy('home')
    
#__________________________________________________________________________________________________________________________
#VISTA BASADA EN CLASES PARA LA VIEW DE LOGIN PERSONALIZADO
#Este login evaluara si es nuevo usuario y si es asi lo va a redireccionar a cambiar la contraseña
@add_group_name_to_context
class CustomLoginView(LoginView):
    def form_invalid(self, form):
        messages.error(self.request, 'Usuario o contraseña incorrectos. Por favor, intente nuevamente.')
        return super().form_invalid(form)

    def form_valid(self, form):
        response = super().form_valid(form)

        # Accedemos al perfil de usuario
        profile = self.request.user.profile

        # Verificamos el valor del campo "creado por el coordinador del modelo Profile"
        if profile.create_by_adr:
            messages.warning(self.request, 'Bienvenido, debe cambiar su contraseña ahora!')
            response['Location'] = reverse_lazy('profile_password_change')
            response.status_code = 302

        return response
    
    def get_success_url(self):
        return super().get_success_url()



# ____________________________________________________________________________________________________________________________
# ____________________________________________________________________________________________________________________________

#VISTA BASADA EN CLASES DE REGISTRO DE USUARIOS (POR EL ADR)

@add_group_name_to_context
class AddUserView(UserPassesTestMixin, LoginRequiredMixin, CreateView):
    model = User
    form_class = UserCreationForm
    template_name = 'profiles/add_user.html'
    success_url = reverse_lazy('profile_list')

    #Creamos una función para que solo el coordinador y el administrador puedan registrar usuarios
    def test_func(self):
        return self.request.user.groups.first().name == 'ADR' 
    
    #Función si el usuario no tiene permiso
    def handle_no_permission(self):
        return redirect('error')


    #Recuperamos los grupos y su singular
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs) #Obtenemos el contexto


        # Obtenemos todos los grupos disponibles, excluyendo los grupos 'Funcionarios' y 'Administradores'
        groups = Group.objects.exclude(name__in=['ADR'])
        singular_groups = [plural_singular(group.name).capitalize() for group in groups] # Obtener los nombres singulares de los grupos
        context['groups'] = zip(groups, singular_groups) #unimos las 2 variables de grupos para obtener el singular del grupo

        return context
    
    #Función para validad y almacenar los datos del formulario
    def form_valid(self, form):

        #Obtener el grupo que se selecciono
        group_id = self.request.POST['group']

        #Obtener el grupo
        group = Group.objects.get(id=group_id)

        #Creamos un usuario sin guardarlo aun
        user = form.save(commit=False)

        #Al nuevo usuario se le asigna la contraseña por defecto "contraseña"
        user.set_password('contraseña')

        # Convertir al usuario a staff solo si el grupo es '2'(ADR)
        if group_id != ['2']:
            user.is_staff = True

        #Creamos al nuevo usuario
        user.save()

        #Limpiamos o eliminamos al usuario que se crea por defecto en el grupo de "Funcionarios"
        user.groups.clear()

        #Agregamos al usuario al grupo seleccionado en el formulario
        user.groups.add(group)

        messages.success(self.request, 'Usuario creado exitosamente')

        return super().form_valid(form)

# ____________________________________________________________________________________________________________________________

#VISTA BASADA EN CLASES PARA CAMBIAR LA CONTRASEÑA DEL USUARIO

@add_group_name_to_context
class ProfilePasswordChangeView(PasswordChangeView):
    template_name = 'profiles/change_password.html'
    success_url = reverse_lazy('home')

    #Definimos la función para realizar el cambio de contraseña
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['password_changed'] = self.request.session.get('password_changed', False) #Obtenemos el valor de la variable de sesión
        return context
    
    #Validamos el formulario de cambio de contraseña
    def form_valid(self, form):

        #Actualizamos el campo "create_by_adr" del modelo Profile
        profile = Profile.objects.get(user=self.request.user)
        profile.create_by_adr = False
        profile.save()


        messages.success(self.request, 'Contraseña cambiada exitosamente')

        #Actualizamos la sesión logeado con la contraseña nueva cambiada
        update_session_auth_hash(self.request, form.user)
        self.request.session['profile_password_changed'] = True
        return super().form_valid(form)
    
    #Si el formulario no es valido
    def form_invalid(self, form):
        messages.error(self.request, 'Las contraseñas no coinciden o no cumple el estándar de seguridad')
        return super().form_invalid(form)


# ____________________________________________________________________________________________________________________________
#VISTA BASADA EN CLASES PARA EDITAR UN USUARIO
@add_group_name_to_context
@add_group_name_to_context
class ProfileUpdateView(LoginRequiredMixin, UpdateView):
    model = Profile
    template_name = 'profiles/profile_edit.html'
    form_class = ProfileForm

    def get_object(self):
        # Recupera el perfil basado en el pk proporcionado en la URL
        profile = get_object_or_404(Profile, pk=self.kwargs['pk'])
        return profile

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        profile = self.get_object()
        user = profile.user  # Obtener el usuario del perfil

        # Recuperamos los forms de user y profile user
        context['user_profile'] = user
        context['profile_form'] = ProfileForm(instance=profile)
        context['user_form'] = UserForm(instance=user)

        # Recuperamos los grupos
        context['singular_groups'] = Group.objects.values_list('name', 'id')
        context['group_id_user'] = user.groups.values_list('id', flat=True).first()
        return context

    def post(self, request, *args, **kwargs):
        profile = self.get_object()
        user = profile.user
        user_form = UserForm(request.POST, instance=user)
        profile_form = ProfileForm(request.POST, request.FILES, instance=profile)

        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()

            messages.success(request, 'Usuario editado exitosamente')
            return redirect('profile_list')

        context = self.get_context_data()
        context['user_form'] = user_form
        context['profile_form'] = profile_form
        return render(request, 'profiles/profile_edit.html', context)

  



#________________________________________________________________________________________________________________________________________
#VISTA BASADA EN CLASES PARA LISTAR LOS USUARIOS REGISTRADOR
@add_group_name_to_context
class ProfileListView(LoginRequiredMixin, ListView):
    model = Profile
    template_name = 'profiles/profile_list.html'
    context_object_name = 'profiles'
    paginate_by = 10 # Número de perfiles por página

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Obtener el grupo del usuario actual
        user = self.request.user
        group_id, group_name, group_name_singular, color = get_group_and_color(user)

        # Agregar información del grupo al contexto
        context['group_name'] = group_name
        context['group_name_singular'] = group_name_singular
        context['color'] = color

        # Agregar nombres de grupos en singular para cada perfil
        profiles_with_singular_groups = []
        for profile in context['profiles']:
            singular_groups = [plural_singular(group.name) for group in profile.user.groups.all()]
            profiles_with_singular_groups.append({
                'profile': profile,
                'singular_groups': singular_groups
            })

        context['profiles_with_singular_groups'] = profiles_with_singular_groups
        
        # Configurar el paginador
        paginator = Paginator(Profile.objects.all(), self.paginate_by)  # Asegúrate de que el queryset sea correcto
        page_number = self.request.GET.get('page')

        try:
            profiles_paginated = paginator.page(page_number)
        except PageNotAnInteger:
            profiles_paginated = paginator.page(1)
        except EmptyPage:
            profiles_paginated = paginator.page(paginator.num_pages)

        context['profiles'] = profiles_paginated

        return context


#__________________________________________________________________________________________________________________________________________
# VISTA BASADA EN CLASES DEL INDEX
class IndexView(TemplateView):
    template_name = 'index.html'
# ____________________________________________________________________________________________________________________________


# VISTA BASADA EN CLASES DEL HOME
@add_group_name_to_context # Decorador
class HomeView(LoginRequiredMixin, TemplateView):
    template_name = 'home.html'

# ____________________________________________________________________________________________________________________________

#VISTA BASADA EN CLASES PARA LA VIEW DE ERROR
@add_group_name_to_context
class ErrorView(TemplateView):
    template_name = 'error.html'

    #FUNCIÓN PARA MOSTRAR LA IMAGEN DE ERROR
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        error_image = os.path.join(settings.MEDIA_ROOT, 'error.png')
        context['error_image'] = error_image
        return context

# ____________________________________________________________________________________________________________________________

#VISTA BASADA EN CLASES PARA LA VIEW DE ALL IN ONE
@add_group_name_to_context  # Decorador
class AllInOneView(LoginRequiredMixin, ListView):
    model = AllInOne  # Importamos el modelo AllInOne
    template_name = 'modulos/all_in_one.html'
    context_object_name = 'all_in_ones'  # Decorador
    paginate_by = 10  # Cantidad de objetos por página

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Obtener el queryset inicial
        queryset = AllInOne.objects.all()

        # Filtrado y paginación usando la función auxiliar
        equipos, filter_ubicacion, ubicaciones = filtrar_y_paginar(
            self.request, queryset, self.paginate_by
        )

        # Agregar los filtros al contexto
        context['all_in_ones'] = equipos
        context['filter_ubicacion'] = filter_ubicacion
        context['ubicaciones'] = ubicaciones

        return context

# ____________________________________________________________________________________________________________________________

#__________________________________________________________________________________________________________________________
#VISTA BASADA EN CLASES PARA AGREGAR UN All IN ONE

@add_group_name_to_context # Decorador
class Add_AllInOneView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = AllInOne  # Importamos el modelo AllInOne 
    template_name = 'modulos/add_all_in_one.html'  

    form_class = AllInOneForm  # Importamos el formulario del modelo AllInOne
    success_url = reverse_lazy('all_in_one')  # Redireccionamos a la lista

    # Función  que verifica si el usuario actual tiene permiso para acceder a esta vista.
    def test_func(self):
        return self.request.user.groups.first().name in ['ADR','Operadores ADR', 'Auxiliares Operadores ADR']

    # Función que se ejecuta cuando un usuario no tiene permiso para acceder a esta vista y lo redirige a una página de error.
    def handle_no_permission(self):
        return redirect('error')

    #Funcion para agregar datos al contexto
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['all_in_one'] = AllInOne.objects.all()
        return context
    
    def form_valid(self, form):
        # Verificar unicidad del número de serie
        n_serie = form.cleaned_data.get('n_serie')
        if AllInOne.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
            # Agregar el error al formulario y redirigir a form_invalid
            form.add_error('n_serie', 'Este Número de Serie ya existe')
            return self.form_invalid(form)

        # Establecer el usuario que está creando el registro
        form.instance.all_one_creado = self.request.user
        messages.success(self.request, 'All In One se ha guardado correctamente.')
        return super().form_valid(form)

    def form_invalid(self, form):
        # Mostrar un mensaje general de error
        messages.error(self.request, 'Por favor, corrija los errores y vuelva a intentarlo.')
        return super().form_invalid(form)
# ____________________________________________________________________________________________________________________________

#VISTA BASADA EN CLASES PARA EDITAR UN All IN ONE
@add_group_name_to_context  # Decorador
class Edit_AllInOneView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = AllInOne  # Importamos el modelo AllInOne   
    template_name = 'modulos/edit_all_in_one.html'
    form_class = AllInOneForm  # Importamos el formulario del modelo AllInOne
    success_url = reverse_lazy('all_in_one')  # Redireccionamos a la lista

    # Función  que verifica si el usuario actual tiene permiso para acceder a esta vista.
    def test_func(self):
        return self.request.user.groups.first().name in ['ADR']

    # Función que se ejecuta cuando un usuario no tiene permiso para acceder a esta vista y lo redirige a una página de error.
    def handle_no_permission(self):
        return redirect('error')
    

    
    def form_valid(self, form):
        # Verificar unicidad del número de serie
        n_serie = form.cleaned_data.get('n_serie')
        if AllInOne.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
            # Agregar el error al formulario y redirigir a form_invalid
            form.add_error('n_serie', 'Este Número de Serie ya existe')
            return self.form_invalid(form)

        form.instance.all_one_creado = self.request.user
        
        messages.success(self.request, 'All In One se ha modificado correctamente.')
        return super().form_valid(form)
    def form_invalid(self, form):
        # Mostrar un mensaje general de error
        messages.error(self.request, 'Por favor, corrija los errores y vuelva a intentarlo.')
        return super().form_invalid(form)
    



# ____________________________________________________________________________________________________________________________
#VISTA BASADA EN CLASES PARA EDITAR UN All IN ONE ADMINISTRADORES

@add_group_name_to_context  # Decorador
class AllInOneAdminView(LoginRequiredMixin, ListView):
    model = AllInOneAdmins
    template_name = 'modulos/all_in_one_adm.html'
    context_object_name = 'all_in_ones'
    paginate_by = 10  # Cantidad de objetos por página

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener los valores únicos de 'ubicacion_all_in_one_admin' para el filtro
        ubicaciones_all_in_one_admin = AllInOneAdmins.objects.values_list('ubicacion_all_in_one_admin', flat=True).distinct()
       
        # Obtener el valor del filtro desde la solicitud GET
        filter_ubicacion_all_in_one_admin = self.request.GET.get('filter_ubicacion_all_in_one_admin', '')

        # Obtener el queryset y aplicar el filtro
        queryset = AllInOneAdmins.objects.all()
        if filter_ubicacion_all_in_one_admin:
            queryset = queryset.filter(ubicacion_all_in_one_admin=filter_ubicacion_all_in_one_admin)

        # Ordenar el queryset
        queryset = queryset.order_by('ubicacion_all_in_one_admin')

        # Configurar el paginador
        paginator = Paginator(queryset, self.paginate_by)
        page_number = self.request.GET.get('page')

        try:
            all_in_ones_paginated = paginator.page(page_number)
        except PageNotAnInteger:
            all_in_ones_paginated = paginator.page(1)
        except EmptyPage:
            all_in_ones_paginated = paginator.page(paginator.num_pages)

        context['all_in_ones'] = all_in_ones_paginated
        context['ubicaciones_all_in_one_admin'] = ubicaciones_all_in_one_admin
        context['filter_ubicacion_all_in_one_admin'] = filter_ubicacion_all_in_one_admin

        return context
    


# ____________________________________________________________________________________________________________________________

#VISTA BASADA EN CLASES PARA AGREGAR UN ALL IN ONE DE UN ADMINISTRADOR

@add_group_name_to_context # Decorador
class Add_AllInOneAdminView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = AllInOneAdmins  # Importamos el modelo AllInOneAdmins   
    template_name ='modulos/add_all_in_one_adm.html'

    form_class = AllInOneAdminsForm  # Importamos el formulario para el modelo AllInOneAdmins
    success_url = reverse_lazy('all_in_one_adm')  # Redireccionamos a la lista

    # Función  que verifica si el usuario actual tiene permiso para acceder a esta vista.
    def test_func(self):
        return self.request.user.groups.first().name in ['ADR','Operadores ADR', 'Auxiliares Operadores ADR']
    
    # Función que se ejecuta cuando un usuario no tiene permiso para acceder a esta vista y lo redirige a una página de error.
    def handle_no_permission(self):
        return redirect('error')
    
    #Funcion para agregar datos al contexto
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['all_in_ones_adms'] = AllInOneAdmins.objects.all()
        return context
    
    def form_valid(self, form):
        # Verificar unicidad del número de serie
        n_serie = form.cleaned_data.get('n_serie')
        if AllInOneAdmins.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
            # Agregar el error al formulario y redirigir a form_invalid
            form.add_error('n_serie', 'Este Número de Serie ya existe')
            return self.form_invalid(form)
    
    # Establecer el usuario que está creando el registro
        form.instance.all_one_adm_creado = self.request.user
        messages.success(self.request, 'All In One Administradores se ha guardado correctamente.')
        return super().form_valid(form)
    
    # Función que se ejecuta cuando el formulario es inválido, muestra un mensaje de error y vuelve a mostrar el formulario.
    def form_invalid(self, form):
        messages.error(self.request, 'Por favor corrija los errores y vuelva a intentarlo.')
        return super().form_invalid(form)


# ____________________________________________________________________________________________________________________________

#VISTA BASADA EN CLASES PARA EDITAR UN All IN ONE ADMINISTRADORES
@add_group_name_to_context  # Decorador
class Edit_AllInOneAdmView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = AllInOneAdmins  # Importamos el modelo AllInOne   
    template_name = 'modulos/edit_all_in_one_adm.html'
    form_class = AllInOneAdminsForm  # Importamos el formulario del modelo AllInOne
    success_url = reverse_lazy('all_in_one_adm')  # Redireccionamos a la lista

    # Función  que verifica si el usuario actual tiene permiso para acceder a esta vista.
    def test_func(self):
        return self.request.user.groups.first().name in ['ADR']

    # Función que se ejecuta cuando un usuario no tiene permiso para acceder a esta vista y lo redirige a una página de error.
    def handle_no_permission(self):
        return redirect('error')
    

    
    def form_valid(self, form):
        # Verificar unicidad del número de serie
        n_serie = form.cleaned_data.get('n_serie')
        if AllInOneAdmins.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
            # Agregar el error al formulario y redirigir a form_invalid
            form.add_error('n_serie', 'Este Número de Serie ya existe')
            return self.form_invalid(form)

        form.instance.all_one_adm_creado = self.request.user
        
        messages.success(self.request, 'Registro All In One Administradores se ha modificado correctamente.')
        return super().form_valid(form)
    def form_invalid(self, form):
        # Mostrar un mensaje general de error
        messages.error(self.request, 'Por favor, corrija los errores y vuelva a intentarlo.')
        return super().form_invalid(form)





# ____________________________________________________________________________________________________________________________
#VISTA BASADA EN CLASES PARA LISTAR LOS NOTEBOOKS

@add_group_name_to_context # Decorador
class NotebooksView(LoginRequiredMixin, ListView):
    model = Notebook
    template_name = 'modulos/notebooks.html'
    context_object_name = 'notebooks'
    paginate_by = 10 # Cantidad de objetos por página

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Obtener los valores únicos de 'ubicacion_notebook' para el filtro
        ubicaciones_notebook = Notebook.objects.values_list('ubicacion_notebook', flat=True).distinct()

        # Obtener el valor del filtro desde la solicitud GET
        filter_ubicacion_notebook = self.request.GET.get('filter_ubicacion_notebook', '')

        # Obtener el queryset y aplicar el filtro
        queryset = Notebook.objects.all()
        if filter_ubicacion_notebook:
            queryset = queryset.filter(ubicacion_notebook=filter_ubicacion_notebook)

        # Ordenar el queryset
        queryset = queryset.order_by('ubicacion_notebook')

        # Configurar el paginador
        paginator = Paginator(queryset, self.paginate_by)
        page_number = self.request.GET.get('page')

        try:
            notebooks_paginated = paginator.page(page_number)
        except PageNotAnInteger:
            notebooks_paginated = paginator.page(1)
        except EmptyPage:
            notebooks_paginated = paginator.page(paginator.num_pages)

        context['notebooks'] = notebooks_paginated
        context['ubicaciones_notebook'] = ubicaciones_notebook
        context['filter_ubicacion_notebook'] = filter_ubicacion_notebook

        return context
        

# ____________________________________________________________________________________________________________________________
#VISTA BASADA EN CLASES PARA AGREGAR UN NOTEBOOK

@add_group_name_to_context # Decorador
class AddNotebooksView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Notebook  # Importamos el modelo Notebook
    template_name ='modulos/add_notebooks.html'

    form_class = NotebooksForm  # Importamos el formulario para el modelo Notebook
    success_url = reverse_lazy('notebooks')  # Redireccionamos a la lista

    # Función  que verifica si el usuario actual tiene permiso para acceder a esta vista.
    def test_func(self):
        return self.request.user.groups.first().name in ['ADR','Operadores ADR', 'Auxiliares Operadores ADR']
    
    # Función que se ejecuta cuando un usuario no tiene permiso para acceder a esta vista y lo redirige a una página de error.
    def handle_no_permission(self):
        return redirect('error')
    
    #Funcion para agregar datos al contexto
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['notebooks'] = Notebook.objects.all()
        return context
    
    def form_valid(self, form):
        # Verificar unicidad del número de serie
        n_serie = form.cleaned_data.get('n_serie')
        if Notebook.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
            # Agregar el error al formulario y redirigir a form_invalid
            form.add_error('n_serie', 'Este Número de Serie ya existe')
            return self.form_invalid(form)

    # Establecer el usuario que está creando el registro
        form.instance.notebook_creado = self.request.user
        messages.success(self.request, 'All In One se ha guardado correctamente.')
        return super().form_valid(form)

    def form_invalid(self, form):
        # Mostrar un mensaje general de error
        messages.error(self.request, 'Por favor, corrija los errores y vuelva a intentarlo.')
        return super().form_invalid(form)
# ____________________________________________________________________________________________________________________________

# ____________________________________________________________________________________________________________________________

#VISTA BASADA EN CLASES PARA EDITAR UN NOTEBOOK
@add_group_name_to_context  # Decorador
class Edit_NotebooksView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Notebook  # Importamos el modelo Notebook   
    template_name = 'modulos/edit_notebook.html'
    form_class = NotebooksForm # Importamos el formulario del modelo Notebook
    success_url = reverse_lazy('notebooks')  # Redireccionamos a la lista

    # Función  que verifica si el usuario actual tiene permiso para acceder a esta vista.
    def test_func(self):
        return self.request.user.groups.first().name in ['ADR']

    # Función que se ejecuta cuando un usuario no tiene permiso para acceder a esta vista y lo redirige a una página de error.
    def handle_no_permission(self):
        return redirect('error')
    
    def form_valid(self, form):
        # Verificar unicidad del número de serie
        n_serie = form.cleaned_data.get('n_serie')
        if Notebook.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
            # Agregar el error al formulario y redirigir a form_invalid
            form.add_error('n_serie', 'Este Número de Serie ya existe')
            return self.form_invalid(form)

        form.instance.notebook_creado = self.request.user
        
        messages.success(self.request, 'Registro Notebook se ha modificado correctamente.')
        return super().form_valid(form)
    def form_invalid(self, form):
        # Mostrar un mensaje general de error
        messages.error(self.request, 'Por favor, corrija los errores y vuelva a intentarlo.')
        return super().form_invalid(form)

# ____________________________________________________________________________________________________________________________  

#VISTA BASA EN CLASES PARA AGREGAR UN MINI PC

@add_group_name_to_context # Decorador
class AddMiniPCView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = MiniPC  # Importamos el modelo MiniPC
    template_name ='modulos/add_mini_pc.html'
    form_class = MiniPCForm  # Importamos el formulario para el modelo MiniPC
    success_url = reverse_lazy('mini_pc')  # Redireccionamos a la lista

    # Función  que verifica si el usuario actual tiene permiso para acceder a esta vista.
    def test_func(self):
        return self.request.user.groups.first().name in ['ADR','Operadores ADR', 'Auxiliares Operadores ADR']
    
    # Función que se ejecuta cuando un usuario no tiene permiso para acceder a esta vista y lo redirige a una página de error.
    def handle_no_permission(self):
        return redirect('error')
    
    #Funcion para agregar datos al contexto
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['mini_pc'] = MiniPC.objects.all()
        return context
    
    def form_valid(self, form):
        # Verificar unicidad del número de serie
        n_serie = form.cleaned_data.get('n_serie')
        if MiniPC.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
            # Agregar el error al formulario y redirigir a form_invalid
            form.add_error('n_serie', 'Este Número de Serie ya existe')
            return self.form_invalid(form)
        
        # Establecer el usuario que está creando el registro
        form.instance.mini_pc_creado = self.request.user
        messages.success(self.request, 'Mini PC se ha guardado correctamente.')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        # Mostrar un mensaje general de error
        messages.error(self.request, 'Por favor, corrija los errores y vuelva a intentarlo.')
        return super().form_invalid(form)

# ____________________________________________________________________________________________________________________________

#VISTA BASADA EN CLASES PARA EDITAR UN MINI PC
@add_group_name_to_context  # Decorador
class Edit_MiniPCView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = MiniPC  # Importamos el modelo Notebook  
    template_name = 'modulos/edit_mini_pc.html'
    form_class = MiniPCForm # Importamos el formulario del modelo Notebook
    success_url = reverse_lazy('mini_pc')  # Redireccionamos a la lista

    # Función  que verifica si el usuario actual tiene permiso para acceder a esta vista.
    def test_func(self):
        return self.request.user.groups.first().name in ['ADR']

    # Función que se ejecuta cuando un usuario no tiene permiso para acceder a esta vista y lo redirige a una página de error.
    def handle_no_permission(self):
        return redirect('error')
    
    def form_valid(self, form):
        # Verificar unicidad del número de serie
        n_serie = form.cleaned_data.get('n_serie')
        if MiniPC.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
            # Agregar el error al formulario y redirigir a form_invalid
            form.add_error('n_serie', 'Este Número de Serie ya existe')
            return self.form_invalid(form)

        form.instance.mini_pc_creado = self.request.user
        
        messages.success(self.request, 'Registro Mini PC se ha modificado correctamente.')
        return super().form_valid(form)
    def form_invalid(self, form):
        # Mostrar un mensaje general de error
        messages.error(self.request, 'Por favor, corrija los errores y vuelva a intentarlo.')
        return super().form_invalid(form)

# ____________________________________________________________________________________________________________________________  


# ____________________________________________________________________________________________________________________________
#VISTA BASADA EN CLASES PARA LISTAR LOS MINI PC

@add_group_name_to_context # Decorador
class MiniPCView(LoginRequiredMixin, ListView):
    model = MiniPC  # Importamos el modelo MiniPC
    template_name = 'modulos/mini_pc.html'
    context_object_name ='minis_pcs' #Decorador
    paginate_by = 10 #Cantidad de registros por página

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Obtener los valores únicos de 'ubicacion_mini_pc' para el filtro
        ubicaciones_mini_pc = MiniPC.objects.values_list('ubicacion_mini_pc', flat=True).distinct()

        # Obtener el valor del filtro desde la solicitud GET
        filter_ubicacion_mini_pc = self.request.GET.get('filter_ubicacion_mini_pc', '')

        # Obtener el queryset y aplicar el filtro
        queryset = MiniPC.objects.all()
        if filter_ubicacion_mini_pc:
            queryset = queryset.filter(ubicacion_mini_pc=filter_ubicacion_mini_pc)

        # Ordenar el queryset
        queryset = queryset.order_by('ubicacion_mini_pc')

        # Configurar el paginador
        paginator = Paginator(queryset, self.paginate_by)
        page_number = self.request.GET.get('page')

        try:
            minis_pcs_paginated = paginator.page(page_number)
        except PageNotAnInteger:
            minis_pcs_paginated = paginator.page(1)
        except EmptyPage:
            minis_pcs_paginated = paginator.page(paginator.num_pages)

        context['minis_pcs'] = minis_pcs_paginated
        context['ubicaciones_mini_pc'] = ubicaciones_mini_pc
        context['filter_ubicacion_mini_pc'] = filter_ubicacion_mini_pc

        return context

# ____________________________________________________________________________________________________________________________
#VISTA BASADA EN CLASES PARA AGREGAR UN PROYECTOR

@add_group_name_to_context
class AddProyectorView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Proyectores
    template_name = 'modulos/add_proyector.html'
    
    form_class = ProyectoresForm # Importamos el formulario para el modelo Proyectores
    success_url = reverse_lazy('proyectores') #Redireccionamos a la lista

    # Función  que verifica si el usuario actual tiene permiso para acceder a esta vista.
    def test_func(self):
        return self.request.user.groups.first().name in ['ADR','Operadores ADR', 'Auxiliares Operadores ADR']
    
    # Función que se ejecuta cuando un usuario no tiene permiso para acceder a esta vista y lo redirige a una página de error.
    def handle_no_permission(self):
        return redirect('error')
    
    #Funcion para agregar datos al contexto
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['proyector'] = Proyectores.objects.all()
        return context
    
    def form_valid(self, form):
        # Verificar unicidad del número de serie
        n_serie = form.cleaned_data.get('n_serie')
        if Proyectores.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
            # Agregar el error al formulario y redirigir a form_invalid
            form.add_error('n_serie', 'Este Número de Serie ya existe')
            return self.form_invalid(form)
        
        # Establecer el usuario que está creando el registro
        form.instance.proyector_creado = self.request.user
        messages.success(self.request, 'El Proyector se ha guardado correctamente.')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        # Mostrar un mensaje general de error
        messages.error(self.request, 'Por favor, corrija los errores y vuelva a intentarlo.')
        return super().form_invalid(form)
# ____________________________________________________________________________________________________________________________

# ____________________________________________________________________________________________________________________________

#VISTA BASADA EN CLASES PARA EDITAR UN PROYECTOR
@add_group_name_to_context  # Decorador
class Edit_ProyectorView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Proyectores
    template_name = 'modulos/edit_proyector.html'
    form_class = ProyectoresForm # Importamos el formulario del modelo Notebook
    success_url = reverse_lazy('proyectores')  # Redireccionamos a la lista

    # Función  que verifica si el usuario actual tiene permiso para acceder a esta vista.
    def test_func(self):
        return self.request.user.groups.first().name in ['ADR']

    # Función que se ejecuta cuando un usuario no tiene permiso para acceder a esta vista y lo redirige a una página de error.
    def handle_no_permission(self):
        return redirect('error')
    
    def form_valid(self, form):
        # Verificar unicidad del número de serie
        n_serie = form.cleaned_data.get('n_serie')
        if Proyectores.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
            # Agregar el error al formulario y redirigir a form_invalid
            form.add_error('n_serie', 'Este Número de Serie ya existe')
            return self.form_invalid(form)

        form.instance.proyector_creado = self.request.user
        
        messages.success(self.request, 'Registro Mini PC se ha modificado correctamente.')
        return super().form_valid(form)
    def form_invalid(self, form):
        # Mostrar un mensaje general de error
        messages.error(self.request, 'Por favor, corrija los errores y vuelva a intentarlo.')
        return super().form_invalid(form)



#VISTA BASADA EN CLASES PARA LISTAR LOS PROYECTORES
@add_group_name_to_context  # Decorador
class ProyectoresView(LoginRequiredMixin, ListView):
    model = Proyectores
    template_name = 'modulos/proyectores.html'
    context_object_name = 'proyectores'  # Decorador
    paginate_by = 10  # Cuantos registros por página

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Obtener los valores únicos de 'ubicacion_proyector' para el filtro
        ubicaciones_proyector = Proyectores.objects.values_list('ubicacion_proyector', flat=True).distinct()

        # Obtener el valor del filtro desde la solicitud GET
        filter_ubicacion_proyector = self.request.GET.get('filter_ubicacion_proyector', '')

        # Obtener el queryset y aplicar el filtro
        queryset = Proyectores.objects.all()
        if filter_ubicacion_proyector:
            queryset = queryset.filter(ubicacion_proyector=filter_ubicacion_proyector)

        # Ordenar el queryset
        queryset = queryset.order_by('ubicacion_proyector')

        # Configurar el paginador
        paginator = Paginator(queryset, self.paginate_by)
        page_number = self.request.GET.get('page')

        try:
            proyectores_paginated = paginator.page(page_number)
        except PageNotAnInteger:
            proyectores_paginated = paginator.page(1)
        except EmptyPage:
            proyectores_paginated = paginator.page(paginator.num_pages)

        context['proyectores'] = proyectores_paginated
        context['ubicaciones_proyector'] = ubicaciones_proyector
        context['filter_ubicacion_proyector'] = filter_ubicacion_proyector

        return context

# ____________________________________________________________________________________________________________________________
#MODULO DE GESTIÓN DE INVENTARIOS
# ____________________________________________________________________________________________________________________________
#VISTA BASADA EN CLASES PARA AGREGAR UN ACTIVO A LA BODEGA ADR
@add_group_name_to_context
class AddBodegaADRView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = BodegaADR
    template_name = 'modulos/add_bodega_adr.html'

    form_class = BodegaADRForm #Importamos el formulario del modelo BodegaADR
    success_url = reverse_lazy('bodega_adr')

    # Función  que verifica si el usuario actual tiene permiso para acceder a esta vista.
    def test_func(self):
        return self.request.user.groups.first().name in ['ADR','Operadores ADR', 'Auxiliares Operadores ADR']

    # Función que se ejecuta cuando un usuario no tiene permiso para acceder a esta vista y lo redirige a una página de error.
    def handle_no_permission(self):
        return redirect('error')

    #Funcion para agregar datos al contexto
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['bodega_adr'] = BodegaADR.objects.all()
        return context
    
    def form_valid(self, form):
        # Verificar unicidad del número de serie
        n_serie = form.cleaned_data.get('n_serie')
        if AllInOne.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
            # Agregar el error al formulario y redirigir a form_invalid
            form.add_error('n_serie', 'Este Número de Serie ya existe')
            return self.form_invalid(form)

        # Establecer el usuario que está creando el registro
        form.instance.bodega_adr_creado = self.request.user
        messages.success(self.request, 'Activo se ha guardado correctamente en Bodega ADR.')
        return super().form_valid(form)

    def form_invalid(self, form):
        # Mostrar un mensaje general de error
        messages.error(self.request, 'Por favor, corrija los errores y vuelva a intentarlo.')
        return super().form_invalid(form)                                                                                                   
# ____________________________________________________________________________________________________________________________

# ____________________________________________________________________________________________________________________________

#VISTA BASADA EN CLASES PARA EDITAR UN ACTIVO EN LA BODEGA ADR
@add_group_name_to_context  # Decorador
class Edit_BodegaADRView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = BodegaADR
    template_name = 'modulos/edit_bodega_adr.html'
    form_class = BodegaADRForm # Importamos el formulario del modelo Bodega ADR
    success_url = reverse_lazy('bodega_adr')  # Redireccionamos a la lista

    # Función  que verifica si el usuario actual tiene permiso para acceder a esta vista.
    def test_func(self):
        return self.request.user.groups.first().name in ['ADR']

    # Función que se ejecuta cuando un usuario no tiene permiso para acceder a esta vista y lo redirige a una página de error.
    def handle_no_permission(self):
        return redirect('error')
    
    def form_valid(self, form):
        # Verificar unicidad del número de serie
        n_serie = form.cleaned_data.get('n_serie')
        if BodegaADR.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
            # Agregar el error al formulario y redirigir a form_invalid
            form.add_error('n_serie', 'Este Número de Serie ya existe')
            return self.form_invalid(form)

        form.instance.bodega_adr_creado = self.request.user
        
        messages.success(self.request, 'Registro Bodega ADR se ha modificado correctamente.')
        return super().form_valid(form)
    def form_invalid(self, form):
        # Mostrar un mensaje general de error
        messages.error(self.request, 'Por favor, corrija los errores y vuelva a intentarlo.')
        return super().form_invalid(form)
#__________________________________________________________________________________________________________________________

#VISTA DE CLASES PARA LISTAR ACTIVOS DE BODEGA ADR

@add_group_name_to_context # Decorador
class BodegaADRView(LoginRequiredMixin, ListView):
    model = BodegaADR
    template_name = 'modulos/bodega_adr.html'
    context_object_name = 'bodegas_adr'
    paginate_by = 10

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener los valores únicos de 'activo' para el filtro
        activos = BodegaADR.objects.values_list('activo', flat=True).distinct()

        # Obtener el valor del filtro desde la solicitud GET
        filter_activo = self.request.GET.get('filter_activo', '')

        # Obtener el queryset y aplicar el filtro
        queryset = BodegaADR.objects.all()
        if filter_activo:
            queryset = queryset.filter(activo=filter_activo)

        # Ordenar el queryset
        queryset = queryset.order_by('activo')
        
        # Configurar el paginador
        paginator = Paginator(queryset, self.paginate_by)
        page_number = self.request.GET.get('page')

        try:
            bodegas_adr_paginated = paginator.page(page_number)
        except PageNotAnInteger:
            bodegas_adr_paginated = paginator.page(1)
        except EmptyPage:
            bodegas_adr_paginated = paginator.page(paginator.num_pages)

        context['bodegas_adr'] = bodegas_adr_paginated
        context['activos'] = activos  # Corrige la variable a 'activos'
        context['filter_activo'] = filter_activo

        return context

# ____________________________________________________________________________________________________________________________

#VISTA BASADA EN CLASES PARA AGREGAR UN ACTIVO A LA AZOTEA DEL ADR

@add_group_name_to_context
class AddAzoteaView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Azotea
    template_name = 'modulos/add_azotea.html'

    form_class = AzoteaForm #Importamos el formulario de Azotea
    success_url = reverse_lazy('azotea_adr')

    # Función  que verifica si el usuario actual tiene permiso para acceder a esta vista.
    def test_func(self):
        return self.request.user.groups.first().name in ['ADR','Operadores ADR', 'Auxiliares Operadores ADR']

    # Función que se ejecuta cuando un usuario no tiene permiso para acceder a esta vista y lo redirige a una página de error.
    def handle_no_permission(self):
        return redirect('error')

    #Funcion para agregar datos al contexto
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['azotea_adr'] = Azotea.objects.all()
        return context
    
    def form_valid(self, form):
        # Verificar unicidad del número de serie
        n_serie = form.cleaned_data.get('n_serie')
        if AllInOne.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
            # Agregar el error al formulario y redirigir a form_invalid
            form.add_error('n_serie', 'Este Número de Serie ya existe')
            return self.form_invalid(form)

        # Establecer el usuario que está creando el registro
        form.instance.azotea_creado = self.request.user
        messages.success(self.request, 'Activo se ha guardado correctamente en Bodega ADR.')
        return super().form_valid(form)

    def form_invalid(self, form):
        # Mostrar un mensaje general de error
        messages.error(self.request, 'Por favor, corrija los errores y vuelva a intentarlo.')
        return super().form_invalid(form)       
    
# ____________________________________________________________________________________________________________________________

#VISTA BASADA EN CLASES PARA EDITAR UN ACTIVO EN LA AZOTEA ADR
@add_group_name_to_context  # Decorador
class Edit_AzoteaView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Azotea
    template_name = 'modulos/edit_azotea_adr.html'
    form_class = AzoteaForm # Importamos el formulario del modelo Azotea ADR
    success_url = reverse_lazy('azotea_adr')  # Redireccionamos a la lista

    # Función  que verifica si el usuario actual tiene permiso para acceder a esta vista.
    def test_func(self):
        return self.request.user.groups.first().name in ['ADR']

    # Función que se ejecuta cuando un usuario no tiene permiso para acceder a esta vista y lo redirige a una página de error.
    def handle_no_permission(self):
        return redirect('error')
    
    def form_valid(self, form):
        # Verificar unicidad del número de serie
        n_serie = form.cleaned_data.get('n_serie')
        if Azotea.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
            # Agregar el error al formulario y redirigir a form_invalid
            form.add_error('n_serie', 'Este Número de Serie ya existe')
            return self.form_invalid(form)

        form.instance.azotea_creado = self.request.user
        
        messages.success(self.request, 'Registro Azotea ADR se ha modificado correctamente.')
        return super().form_valid(form)
    def form_invalid(self, form):
        # Mostrar un mensaje general de error
        messages.error(self.request, 'Por favor, corrija los errores y vuelva a intentarlo.')
        return super().form_invalid(form)
#__________________________________________________________________________________________________________________________

# ____________________________________________________________________________________________________________________________

#VISTA BASADA EN CLASES PARA LISTAR LOS ACTIVOS DE LA AZOTEA
@add_group_name_to_context # Decorador
class AzoteaView(LoginRequiredMixin, ListView):
    model = Azotea  # Importamos el modelo Azotea
    template_name = 'modulos/azotea_adr.html'
    context_object_name = 'azoteas_adr' #Decorador
    paginate_by = 10

    def get_context_data(self, **kwargs): #Decorador
        context = super().get_context_data(**kwargs) # Recuperamos los datos del contexto 

        # Obtener los valores únicos de 'activo' para el filtro
        activos = Azotea.objects.values_list('activo', flat=True).distinct()

        # Obtener el valor del filtro desde la solicitud GET
        filter_activo = self.request.GET.get('filter_activo', '')

        # Obtener el queryset y aplicar el filtro
        queryset = Azotea.objects.all()
        if filter_activo:
            queryset = queryset.filter(activo=filter_activo)

        # Ordenar el queryset
        queryset = queryset.order_by('activo')

        # Configurar el paginador
        paginator = Paginator(queryset, self.paginate_by)
        page_number = self.request.GET.get('page')

        try:
            azoteas_adr_paginated = paginator.page(page_number)
        except PageNotAnInteger:
            azoteas_adr_paginated = paginator.page(1)
        except EmptyPage:
            azoteas_adr_paginated = paginator.page(paginator.num_pages)

        context['azoteas_adr'] = azoteas_adr_paginated
        context['activos'] = activos
        context['filter_activo'] = filter_activo
        
        return context

# ____________________________________________________________________________________________________________________________


#------------------------------------------#
#MODULOS PARA GESTION DE EQUIPOS EN REPORTE
#------------------------------------------#

# ____________________________________________________________________________________________________________________________
#VISTA BASADA EN CLASES PARA AGREGAR REGISTRO DE UN EQUIPO RECEPCIÓN DE EQUIPO A REPORTAR
@add_group_name_to_context
class AddReporteView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Reporte
    template_name = 'modulos/add_reporte_equipos.html'

    form_class = ReporteForm #Importamos el formulario de Reporte
    success_url = reverse_lazy('reporte_equipos')

    # Función  que verifica si el usuario actual tiene permiso para acceder a esta vista.
    def test_func(self):
        return self.request.user.groups.first().name in ['ADR','Operadores ADR', 'Auxiliares Operadores ADR']
    
    # Función que se ejecuta cuando un usuario no tiene permiso para acceder a esta vista y lo redirige a una página de error.
    def handle_no_permission(self):
        return redirect('error')
    
    #Funcion para agregar datos al contexto
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['reporte'] = Reporte.objects.all()
        return context
    
    def form_valid(self, form):
        # Verificar unicidad del número de serie
        n_serie = form.cleaned_data.get('n_serie')
        if Reporte.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
            # Agregar el error al formulario y redirigir a form_invalid
            form.add_error('n_serie', 'Este Número de Serie ya existe')
            return self.form_invalid(form)
        
        # Establecer el usuario que está creando el registro
        form.instance.reporte_creado = self.request.user
        messages.success(self.request, 'Reporte se ha guardado correctamente.')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        # Mostrar un mensaje general de error
        messages.error(self.request, 'Por favor, corrija los errores y vuelva a intentarlo.')
        return super().form_invalid(form)
    

    
# ____________________________________________________________________________________________________________________________
#VISTA BASADA EN CLASES PARA LISTAR LOS EQUIPOS REPORTADOS


@add_group_name_to_context
class ReporteView(LoginRequiredMixin, ListView):
    model = Reporte
    template_name = 'modulos/reporte_equipos.html'
    context_object_name = 'reporte_equipos'
    paginate_by = 1 # Define el número de objetos por página

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Obtener los valores únicos de 'activo' para el filtro
        activos = Reporte.objects.values_list('activo', flat=True).distinct()
        
        # Obtener el valor del filtro desde la solicitud GET
        filter_activo = self.request.GET.get('filter_activo', '')

        # Obtener el queryset y aplicar el filtro
        queryset = Reporte.objects.all()
        if filter_activo:
            queryset = queryset.filter(activo=filter_activo)
        
        # Ordenar el queryset en orden descendente por el campo 'activo'
        queryset = queryset.order_by('activo')

        # Configurar el paginador
        paginator = Paginator(queryset, self.paginate_by)
        page_number = self.request.GET.get('page')

        try:
            reporte_paginated = paginator.page(page_number)
        except PageNotAnInteger:
            reporte_paginated = paginator.page(1)
        except EmptyPage:
            reporte_paginated = paginator.page(paginator.num_pages)

        context['reporte_equipos'] = reporte_paginated
        context['activos'] = activos
        context['filter_activo'] = filter_activo

        return context

        


# ___________________________________________________________________________________________________________________________

#VISTA BASADA EN CLASES PARA DETALLE DE UN EQUIPO REPORTADO

class ReporteEquipoDetailView(LoginRequiredMixin, DetailView):
    model = Reporte
    template_name = 'modulos/detail_reporte_equipo.html'
    form_class = ReporteForm
    context_object_name = 'reporte_equipo'

    # Función  que verifica si el usuario actual tiene permiso para acceder a esta vista.
    def test_func(self):
        return self.request.user.groups.first().name in ['ADR','Operadores ADR', 'Auxiliares Operadores ADR']

    # Función que se ejecuta cuando un usuario no tiene permiso para acceder a esta vista y lo redirige a una página de error.
    def handle_no_permission(self):
        return redirect('error')

# ____________________________________________________________________________________________________________________________
#VISTA BASADA EN CLASES PARA ACTUALIZAR O HACER UPDATE AL REGISTRO DE UN EQUIPO REPORTADO

@add_group_name_to_context
class EditReporteEquipoView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Reporte
    template_name = 'modulos/edit_reporte_equipos.html'
    form_class = ReporteForm
    success_url = reverse_lazy('reporte_equipos')

    # Función  que verifica si el usuario actual tiene permiso para acceder a esta vista.
    def test_func(self):
        return self.request.user.groups.first().name in ['ADR','Operadores ADR', 'Auxiliares Operadores ADR']

    # Función que se ejecuta cuando un usuario no tiene permiso para acceder a esta vista y lo redirige a una página de error.
    def handle_no_permission(self):
        return redirect('error')
    
    def form_valid(self, form):
        # Verificar unicidad del número de serie
        n_serie = form.cleaned_data.get('n_serie')
        if Reporte.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
            # Agregar el error al formulario y redirigir a form_invalid
            form.add_error('n_serie', 'Este Número de Serie ya existe')
            return self.form_invalid(form)
        
        # Si se marca como reportandose, poner en_reporte en False
         # Si se marca como entregado, poner en_reporte en False
        if form.cleaned_data.get('en_reporte'):
            form.instance.entregado = False

        # Establecer el usuario que está creando el registro
        form.instance.reporte_creado = self.request.user
        messages.success(self.request, 'Reporte se ha guardado correctamente.')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        # Mostrar un mensaje general de error
        messages.error(self.request, 'Por favor, corrija los errores y vuelva a intentarlo.')
        return super().form_invalid(form)
    
#___________________________________________________________________________________________________________________________________________________________________________________

#VISTA BASADA EN CLASES PARA LISTAR LOS EQUIPOS QUE SE ENTREGAN
@add_group_name_to_context
class EntregaEquiposView(LoginRequiredMixin, ListView):
    model = Reporte
    template_name = 'modulos/entrega_equipos_reporte.html'
    context_object_name = 'entrega_equipos'
    paginate_by =10  # Número de objetos por página, ajusta según sea necesario

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Obtener los valores únicos de 'activo' para el filtro
        activos = Reporte.objects.values_list('activo', flat=True).distinct()
        
        # Obtener el valor del filtro desde la solicitud GET
        filter_activo = self.request.GET.get('filter_activo', '')

        # Obtener el queryset y aplicar el filtro
        queryset = Reporte.objects.all()
        if filter_activo and filter_activo != '':
            queryset = queryset.filter(activo=filter_activo)
        
        # Ordenar el queryset en orden ascendente por el campo 'activo'
        queryset = queryset.order_by('activo')

        # Paginación
        paginator = Paginator(queryset, self.paginate_by)
        page = self.request.GET.get('page')
        try:
            entrega_equipos = paginator.page(page)
        except PageNotAnInteger:
            # Si la página no es un entero, muestra la primera página.
            entrega_equipos = paginator.page(1)
        except EmptyPage:
            # Si la página está fuera del rango (p.ej. 9999), muestra la última página de resultados.
            entrega_equipos = paginator.page(paginator.num_pages)

        context['entrega_equipos'] = entrega_equipos
        context['activos'] = activos
        context['filter_activo'] = filter_activo

        return context

    
    #______________________________________________________________________________________________________________________________

    #VISTA BASADA EN CLASES PARA ACTUALIZAR O HACER UPDATE AL REGISTRO DE UN EQUIPO REPORTADO

@add_group_name_to_context
class EditEntregaEquipoView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Reporte
    template_name = 'modulos/edit_entrega_equipo.html'
    form_class = ReporteForm
    success_url = reverse_lazy('entrega_equipos_reporte')

    # Función  que verifica si el usuario actual tiene permiso para acceder a esta vista.
    def test_func(self):
        return self.request.user.groups.first().name in ['ADR','Operadores ADR', 'Auxiliares Operadores ADR']

    # Función que se ejecuta cuando un usuario no tiene permiso para acceder a esta vista y lo redirige a una página de error.
    def handle_no_permission(self):
        return redirect('error')
    
    def form_valid(self, form):
        # Verificar unicidad del número de serie
        n_serie = form.cleaned_data.get('n_serie')
        if Reporte.objects.exclude(id=form.instance.id).filter(n_serie=n_serie).exists():
            # Agregar el error al formulario y redirigir a form_invalid
            form.add_error('n_serie', 'Este Número de Serie ya existe')
            return self.form_invalid(form)
        

        # Si se marca como entregado, poner en_reporte en False
        if form.cleaned_data.get('entregado'):
            form.instance.en_reporte = False
        
        # Establecer el usuario que está creando el registro
        form.instance.operador_entrega = self.request.user
        messages.success(self.request, 'Reporte se ha guardado correctamente.')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        # Mostrar un mensaje general de error
        messages.error(self.request, 'Por favor, corrija los errores y vuelva a intentarlo.')
        return super().form_invalid(form)
    